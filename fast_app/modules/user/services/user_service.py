from collections import defaultdict
from io import BytesIO
from typing import Any, Optional, Dict, List, Tuple
from datetime import datetime

from beanie import PydanticObjectId

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from beanie.operators import In
from fast_app.utils.logger import logger
from fast_app.defaults.common_enums import UserRole
from fast_app.defaults.user_enums import UserActivityStatusEnum
from fast_app.modules.product.models.product_model import Product
from fast_app.modules.user.models.user_device_model import UserDevice
from fast_app.modules.user.models.user_model import User
from fast_app.modules.user.schemas.user_schema import (
    UpdateAdminPermissions,
    UserCreateForm,
    UserUpdateForm,
)
from fast_app.utils.common_utils import escape_regex, exclude_unset, stringify_object_ids
from fast_app.utils.file_utils import upload_files
from fast_app.utils.logger import logger

# -----------------------------------------------------
# LIST USERS
# -----------------------------------------------------
async def get_users(
    page: int = 1,
    limit: int = 10,
    search: Optional[str] = None,
    sort: Optional[str] = None,
    filters: Optional[Dict] = None,
) -> Tuple[List[dict], Dict[str, Any]]:

    pipeline = []
    match_stage: Dict[str, Any] = {"is_deleted": False}

    if search:
        safe_search=escape_regex(search)
        match_stage["$or"] = [
            {"full_name": {"$regex": safe_search, "$options": "i"}},
            {"email": {"$regex": safe_search, "$options": "i"}},
        ]

    if filters:
        if "status" in filters:
            match_stage["status"] = filters["status"]
        if "role" in filters:
            match_stage["role"] = filters["role"]
        if "created_at" in filters:
            match_stage["created_at"] = filters["created_at"]

    pipeline.append({"$match": match_stage})

    sort_field = sort.lstrip("-") if sort else "created_at"
    sort_dir = -1 if sort and sort.startswith("-") else 1

    users, pagination = await User.aggregate_with_pagination(
        pipeline=pipeline,
        page=page,
        limit=limit,
        sort_field=sort_field,
        sort_dir=sort_dir,
    )

    return (
        [
            User.model_validate(user).model_dump(by_alias=True, mode="json")
            for user in users
        ],
        pagination,
    )


# -----------------------------------------------------
# Export data
# -----------------------------------------------------
async def export_users(
    search: Optional[str] = None,
    sort: Optional[str] = None,
    filters: Optional[Dict] = None,
) -> List[dict]:

    pipeline = []
    match_stage: Dict[str, Any] = {"is_deleted": False}

    if search:
        safe_search=escape_regex(search)
        match_stage["$or"] = [
            {"full_name": {"$regex": safe_search, "$options": "i"}},
            {"email": {"$regex": safe_search, "$options": "i"}},
        ]

    if filters:
        if "status" in filters:
            match_stage["status"] = filters["status"]
        if "role" in filters:
            match_stage["role"] = filters["role"]
        if "created_at" in filters:
            match_stage["created_at"] = filters["created_at"]

    pipeline.append({"$match": match_stage})

    sort_field = sort.lstrip("-") if sort else "created_at"
    sort_dir = -1 if sort and sort.startswith("-") else 1
    
    pipeline.append({"$sort": {sort_field: sort_dir}})

    users = await User.aggregate_list(
        pipeline=pipeline,
    )

    return [
            User.model_validate(user).model_dump(by_alias=True, mode="json")
            for user in users
        ]


def format_date(value):
    if isinstance(value, str):
        return value.replace("T", " ").split(".")[0]
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return ""


def autosize(ws):
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 22


def style_header(ws):
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    ws.row_dimensions[1].height = 26



def generate_users_excel(users: list, role: UserRole | None = None) -> BytesIO:
    wb = Workbook()
    if wb.active is not None:
        wb.remove(wb.active)

    # ---------------- sub-constructor/end_users ----------------
    if role in (None, UserRole.END_USER):
        ws_user = wb.create_sheet("Sub Constructors")
        ws_user.append([
            "Full Name",
            "Email",
            "Phone Number",
            "City",
            "State",
            "Status",
            "Registered At",
        ])

        style_header(ws_user)

        for u in users:
            if u.get("role") == UserRole.END_USER:
                geo = u.get("geo_location") or {}
                ws_user.append([
                    u.get("full_name", ""),
                    u.get("email", ""),
                    u.get("phone_number") or "",
                    geo.get("city", ""),
                    geo.get("state", ""),
                    u.get("status", ""),
                    format_date(u.get("created_at")),
                ])

        autosize(ws_user)

    # ---------------- store-owner/vendor ----------------
    elif role in (UserRole.VENDOR):
        ws_vendor = wb.create_sheet("Store Owners")
        ws_vendor.append([
            "Full Name",
            "Business Name",
            "Business Email",
            "Email",
            "Phone Number",
            "GST Number",
            "License Number",
            "Status",
            "Registered At",
        ])

        style_header(ws_vendor)

        for u in users:
            if u.get("role") == UserRole.VENDOR:
                ws_vendor.append([
                    u.get("full_name", ""),
                    u.get("business_name", ""),
                    u.get("business_email", ""),
                    u.get("email", ""),
                    u.get("phone_number") or "",
                    u.get("gst_number", ""),
                    u.get("lisence_number", ""),
                    u.get("status", ""),
                    format_date(u.get("created_at")),
                ])

        autosize(ws_vendor)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


# -----------------------------------------------------
# GET USER
# -----------------------------------------------------
async def get_user_by_id(user_id: str):
    try:
        pipeline = [
            {
                "$match": {
                    "_id": PydanticObjectId(user_id),
                    "is_deleted": False,
                }
            },
            {
                "$lookup": {
                    "from": Product.Settings.name,  # usually "products"
                    "localField": "products",
                    "foreignField": "_id",
                    "as": "products",
                }
            },
        ]
        
        coll = User.get_pymongo_collection()
        cursor = coll.aggregate(pipeline)
        result = await cursor.to_list(length=1)  # type: ignore
        
        if not result:
            return None

        return stringify_object_ids(result[0])

    except Exception as e:
        logger.error(str(e))
        return None



# -----------------------------------------------------
# CREATE USER
# -----------------------------------------------------
async def create_new_user(user_data: UserCreateForm):
    if await User.find_one(User.email == user_data.email, User.is_deleted == False):
        raise ValueError("Email already registered")
    
    if user_data.profile_image:
        upload_result = await upload_files([user_data.profile_image], "profile-images")
        image_data = upload_result[0]

    user = User(
        first_name=user_data.first_name,
        last_name=user_data.last_name,
        email=user_data.email,
        role=user_data.role,
        profile_image=image_data.get("path",""),
        password=user_data.password,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
    )

    await user.create()
    return user.model_dump(by_alias=True, mode="json")


# -----------------------------------------------------
# UPDATE USER
# -----------------------------------------------------
async def update_user_details(user_id: str, user_data: UserUpdateForm):
    update_data = user_data.model_dump(exclude_unset=True)

    if not update_data:
        return None

    if user_data.profile_image:
        upload_result = await upload_files([user_data.profile_image], "profile-images")
        image_data = upload_result[0]
        update_data["profile_image"]=image_data.get("path","")
    
    
    user = await User.get(PydanticObjectId(user_id))
    if not user or user.is_deleted:
        return None
    
    update_data["full_name"]=f"{user_data.first_name or user.first_name} {user_data.last_name or user.last_name}"

    update_data["updated_at"] = datetime.utcnow()
    await user.set(exclude_unset(update_data))

    return user.model_dump(by_alias=True, mode="json")


async def update_user_permissions(user_id: str, user_data: UpdateAdminPermissions):
    update_data = user_data.model_dump(exclude_unset=True)
    if not update_data:
        return None

    user = await User.get(PydanticObjectId(user_id))
    if not user or user.is_deleted or user.role is not UserRole.ADMIN:
        return None

    update_data["updated_at"] = datetime.utcnow()
    await user.set(update_data)

    return user.model_dump(by_alias=True, mode="json")


# -----------------------------------------------------
# UPDATE STATUS
# -----------------------------------------------------
async def change_user_status(user_id: str, status):
    user = await User.get(PydanticObjectId(user_id))
    if not user or user.is_deleted:
        return None

    await user.set(
        {
            "status": status,
            "updated_at": datetime.utcnow(),
        }
    )

    return user.model_dump(by_alias=True, mode="json")


# -----------------------------------------------------
# SOFT DELETE
# -----------------------------------------------------
async def remove_user(user_id: str) -> bool:
    user = await User.get(PydanticObjectId(user_id))
    if not user or user.is_deleted:
        return False

    await user.set(
        {
            "is_deleted": True,
            "updated_at": datetime.utcnow(),
        }
    )
    return True


def get_status_room_by_user_id(user_id: str) -> str:
    return f"user_status_room_{user_id}"


async def update_user_activity_status(access_token: str, status: UserActivityStatusEnum):
    try:
        user_device=await UserDevice.find_one(UserDevice.access_token==access_token)
        if user_device is not None:
            await user_device.set({
                "current_status":status,
                "last_active":datetime.now() if status==UserActivityStatusEnum.OFFLINE else None
            })
            logger.info("User status update successfully")
            return True
        logger.warning("User not found with the given token, status update failed")
        return False
    except Exception as e:
        logger.error(e)
        return False


async def get_status_by_user_ids(id_list: List[str]) -> Dict[str, dict]:
    user_ids = [PydanticObjectId(uid) for uid in id_list]

    devices = await UserDevice.find(
        In(UserDevice.user_id, user_ids),
        UserDevice.is_deleted == False,
        UserDevice.expired == False,
    ).to_list()

    devices_by_user = defaultdict(list)
    for device in devices:
        devices_by_user[str(device.user_id)].append(device)

    result: Dict[str, dict] = {}

    for uid in id_list:
        user_devices = devices_by_user.get(uid, [])

        is_online = any(
            d.device_token
            and d.current_status == UserActivityStatusEnum.ONLINE
            for d in user_devices
        )

        if is_online:
            result[uid] = {
                "activity_status": UserActivityStatusEnum.ONLINE,
                "last_seen": None,
            }
        else:
            last_seen = max(
                (d.last_active for d in user_devices if d.last_active),
                default=None,
            )

            result[uid] = {
                "activity_status": UserActivityStatusEnum.OFFLINE,
                "last_seen": last_seen.isoformat() if last_seen else None,
            }

    return result